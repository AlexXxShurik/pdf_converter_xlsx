"""Мультилистовой writer AccKee.

Группировка таблиц в листы по «надёжному» набору заголовков: Dc, Ds,
Specification/Item, Store, Applicable Insert, kg и L (L учитывается только
если распознан как 'L'; l1/l2/l3 — ненадёжные, не сравниваются). Spare Parts
идёт в отдельный лист.
"""
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

NUMBER_RE = re.compile(r"^\d+(?:\.\d+)?$")
RELIABLE = {'Dc', 'Ds', 'Specification/Item', 'Store', 'Applicable Insert',
            'kg', 'L'}
UNRELIABLE = {'l1', 'l2', 'l3', '', '-', 'None', 'q2', 'q3', '1', '2', '3'}


def _reliable_set(labels):
    """Надёжные метки таблицы (для сравнения наборов заголовков).

    l1/l2/l3, пустые и '-' — ненадёжны всегда. L надёжен только если
    распознан точно как 'L'.
    """
    s = set()
    for lab in labels:
        lab = (lab or '').strip()
        if lab in ('l1', 'l2', 'l3') or lab in ('', '-', 'None'):
            continue
        if lab == 'L':
            s.add('L')
        elif lab in RELIABLE:
            s.add(lab)
    return frozenset(s)


def _canonical(lab):
    """Метка каноническая (надёжная для сравнения) и её эталонное значение."""
    lab = (lab or '').strip()
    if lab == 'L':
        return 'L'
    if lab in RELIABLE:
        return lab
    return None


def _compatible(a_labels, b_labels):
    """Два набора заголовков совместимы (один лист)?

    Правило: общее количество колонок совпадает И по значению и порядку
    совпало >= половины позиций → таблицы одинаковые. Сверяются только
    канонические метки (Dc, Ds, Spec, Store, L, AI, kg); l1/l2/l3 и
    мусорные OCR-метки ('Fig.', '8', 'Fi') не учитываются — на них
    распознавание ненадёжно.
    """
    if len(a_labels) != len(b_labels):
        return False
    n = len(a_labels)
    if n == 0:
        return True
    compared = 0
    matches = 0
    for aj, bj in zip(a_labels, b_labels):
        ca, cb = _canonical(aj), _canonical(bj)
        if ca is None or cb is None:
            continue
        compared += 1
        if ca == cb:
            matches += 1
    if compared == 0:
        return True
    return matches >= (compared + 1) // 2


def to_cell_value(text):
    """Значение ячейки: числа → float, остальное (/, SP.., ●, ●) → строка."""
    if text is None:
        return None
    if isinstance(text, (int, float)):
        return text
    text = str(text).strip()
    if not text:
        return None
    if NUMBER_RE.match(text):
        return float(text)
    return text


def _sheet_name(table):
    labels = [c.get('label') for c in table['cols']]
    if any(l in ('type', 'diameter', 'Blade model') for l in labels):
        return 'Spare Parts'
    if len(labels) == 3 and 'Dc' in labels:
        return 'C-KSD Series'
    if 'Dc' in labels or 'Dc' in _reliable_set(labels):
        return 'Drilling'
    return 'Sheet'


def group_tables(tables):
    """Разложить таблицы по листам (совместимые наборы заголовков)."""
    groups = []  # [{'name', 'tables': [...]}]
    for t in tables:
        labels = [c.get('label') for c in t['cols']]
        g = None
        for cand in groups:
            if _compatible(cand['labels'], labels):
                g = cand
                break
        if g is None:
            g = {'name': _sheet_name(t), 'labels': labels, 'tables': []}
            groups.append(g)
        g['tables'].append(t)
    return groups


def _merge_header_labels(ws, col_start, labels):
    """Слить повторяющиеся соседние метки заголовка (напр. Screw×3)."""
    n = len(labels)
    row = 1
    for i in range(n):
        ws.cell(row=row, column=col_start + i, value=labels[i])
        ws.cell(row=row, column=col_start + i).font = Font(bold=True)
    i = 0
    while i < n:
        j = i
        while j + 1 < n and labels[j + 1] == labels[i]:
            j += 1
        if j > i:
            ws.merge_cells(
                start_row=row, start_column=col_start + i,
                end_row=row, end_column=col_start + j)
        i = j + 1


def _best_labels(tables):
    """Метки заголовков группы: по каждой колонке берём самую надёжную.

    Предпочитаем канонические метки (Dc, Ds, ..., L) вместо '-', 'Q2' и пр.
    """
    n = len(tables[0]['cols'])
    best = [None] * n
    for t in tables:
        for j, c in enumerate(t['cols']):
            lab = (c.get('label') or '').strip()
            if not lab:
                continue
            if lab in RELIABLE or lab in ('l1', 'l2', 'l3'):
                best[j] = lab
            elif best[j] is None:
                best[j] = lab
    return [b if b is not None else '' for b in best]


def write_xlsx(tables, path):
    """Записать список таблиц (parse_page) в мультилистовой XLSX."""
    groups = group_tables(tables)
    wb = Workbook()
    wb.remove(wb.active)
    for g in groups:
        ws = wb.create_sheet(title=g['name'])
        labels = _best_labels(g['tables'])
        _merge_header_labels(ws, 1, labels)
        for t in g['tables']:
            for row in t['rows']:
                ws.append([to_cell_value(v) for v in row])
        for col in range(1, len(labels) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
    wb.save(path)
    return path
